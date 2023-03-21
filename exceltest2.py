from openpyxl import load_workbook
wb = load_workbook('testexcel.xlsx')

ws = wb.active
# ws = wb["Sheet2"]

numberOfRows = 6
numberOfColumns = 2

# print(ws["B5"].value)
# print(ws.cell(row=a, column=2).value)

for currentColumn in range(1, numberOfColumns+1):
	print("The values in column " + str(currentColumn) + ":")
	for currentRow in range(1, numberOfRows+1):
		print(ws.cell(row=currentRow, column=currentColumn).value)





# print(wb2.sheetnames)